#!/usr/bin/env python
# coding: utf-8

# In[48]:


import requests
import json
import pygsheets
import re
from transliterate import translit, get_available_language_codes
from openpyxl import Workbook
from requests.exceptions import HTTPError
from fuzzywuzzy import fuzz


# In[49]:


# groups id
group57_morning=142
group57_evening=143
group89_morning=144
group89_evening=145
group89_march_morning=156
group89_march_evening=155

TOKEN = 'c761d6331014ee2e925b6d8606a679d8868dcaac52004d10b0bbbcd98af9315ff1d51c8df9bcb0ce19bfad7ebf395851'


# In[50]:


def get_contest_ids():
    try:
        resp57 = requests.get('https://contest.misis.ru/api/contest/all?token='+TOKEN, 
                              params={'count':1000, 'query':'5-7'})
        resp57.raise_for_status()
        resp89 = requests.get('https://contest.misis.ru/api/contest/all?token='+TOKEN, 
                              params={'count':1000, 'query':'8-9'})
        resp89.raise_for_status()
    except HTTPError as http_err:
        print(f'HTTP error occurred: {http_err}')
    except Exception as err:
        print(f'Error occurred: {err}')
    
    resp57 = resp57.json()
    resp89 = resp89.json()
    contest_ids_57 = [item['id'] for item in resp57['contests']]
    contest_ids_89 = []
    contest_ids_march89 = []
    for item in resp89['contests']:
        groups = {gr['id'] for gr in item['allowedGroups']}
        if group89_morning in groups or group89_evening in groups:
            contest_ids_89.append(item['id'])
        else:
            contest_ids_march89.append(item['id'])
            
    return contest_ids_57, contest_ids_89, contest_ids_march89


# In[51]:


def get_rating():
    contest_ids_57, contest_ids_89, contest_ids_march89 = get_contest_ids()
    
    try:
        rating57 = requests.post('https://contest.misis.ru/api/admin/ratingTable?token='+TOKEN, 
                                 data={'contestIds': contest_ids_57, 'scoreInPractice': 1, 'scoreInTime': 1})
        rating57.raise_for_status()
        rating89 = requests.post('https://contest.misis.ru/api/admin/ratingTable?token='+TOKEN,
                           data={'contestIds': contest_ids_89, 'scoreInPractice': 1, 'scoreInTime': 1})
        rating89.raise_for_status()
        rating_march89 = requests.post('https://contest.misis.ru/api/admin/ratingTable?token='+TOKEN,
                           data={'contestIds': contest_ids_march89, 'scoreInPractice': 1, 'scoreInTime': 1})
        rating_march89.raise_for_status()
    except HTTPError as http_err:
        print(f'HTTP error occurred: {http_err}')
    except Exception as err:
        print(f'Error occurred: {err}')
        
    rating57 = rating57.json()
    rating89 = rating89.json()
    rating_march89 = rating_march89.json()
        
    return rating57, rating_march89


# In[52]:


def get_kids_data():
    try:
        gc = pygsheets.authorize()
        sh = gc.open('Кружок NlogN (Ответы)')
        wks = sh.sheet1
        kids = wks.get_all_records()
    except pygsheets.AuthenticationError as err:
        print(f'Authentication Error occurred: {err}')
    except pygsheets.SpreadsheetNotFound as err:
        print(f'Spreadsheet Not Found: {err}')
    except pygsheets.WorksheetNotFound as err:
        print(f'Worksheet Not Found: {err}')
    except Exception as err:
        print(f'Error occurred: {err}')
    
    # удаление буквы класса
    for kid in kids:
        if type(kid['Класс']) is int:
            continue
        nums = re.findall(r'\d+', kid['Класс'])
        if len(nums) > 0:
            kid['Класс'] = nums[0]
    return kids


# In[87]:


def check(name1, name2):
    n = len(name1)
    m = len(name2)
    dp = [[0] * (m + 1)] * (n + 1)
    #
    for i in range(1, n + 1):
        for j in range(1, m + 1):
            if name1[i - 1] == name2[j - 1]:
                dp[i][j] = max(dp[i][j], dp[i-1][j-1] + 1)
            dp[i][j] = max(dp[i][j], max(dp[i-1][j], dp[i][j-1]))
    if dp[n][m] > max(n, m) * 0.7:
        return True;
    return False;


# In[95]:


def merge(rating_list, info_list):
    for person in rating_list:
        if re.search(r'[а-яА-Я]',person['name']):
            ru_name = person['name'].lower()
            en_name = translit(person['name'], 'ru', reversed = True).lower()
        else:
            en_name = person['name'].lower()
            ru_name = translit(person['name'], 'ru').lower()
        for kid in info_list:
            kids_name = kid['name']
            reversed_name = kid['reversed_name']
            if (check(ru_name, kids_name) or check(en_name, kids_name) or
                check(ru_name, reversed_name) or check(en_name, reversed_name)):
                for key, value in kid.items():
                    if key != 'name':
                        person[key] = value
                break
    return rating_list


# In[109]:


def split_classes(kids, grades):
    splited_list = {'rest' : []}
    for kid in kids:
        try:
            if not kid['grade'] in splited_list:
                splited_list[kid['grade']] = []
            if not int(kid['grade']) in grades:
                 splited_list['rest'].append(kid)
            else:
                splited_list[kid['grade']].append(kid)
        except:
            splited_list['rest'].append(kid)
    return splited_list


# In[135]:


def write_to_sheet(sheet, list_info, important_staff, head_line):
    if sheet.max_row == 1:
        sheet.append(head_line)
    widhts = [0]*len(head_line)
    for i in range(len(head_line)):
        widhts[i] = len(str(head_line))
    for i in range(len(list_info)):
        kid = []
        for item in important_staff:
            try:
                kid.append(list_info[i][item])
            except:
                kid.append(' ')
        sheet.append(kid)
        for j in range(len(kid)):
            widhts[j] = max(widhts[j], len(str(kid[j])))
    for i in range(len(widhts)):      
        sheet.column_dimensions[chr(ord('A') + i)].width = widhts[i]/4


# In[ ]:





# In[123]:


kids_data = get_kids_data() 


# In[124]:


kids_info = []
for kid in kids_data:
    person = {}
    person['name'] = str(kid['Имя']) + ' ' + str(kid['Фамилия'])
    person['reversed_name'] = str(kid['Фамилия']) + ' ' + str(kid['Имя'])
    person['grade'] = kid['Класс']
    person['mail'] = kid['E-mail']
    person['phone_number'] = kid['Телефон, к которому привязан ваш telegram']
    person['telegram'] = kid['Имя пользователя в telegram']
    person['city'] = kid["Город"]
    person['school']=kid ["Школа"]
    kids_info.append(person)


# In[125]:


rating57_row, rating_89_row = get_rating()
rating_57 = []
rating_89 = []

for item in rating57_row['rows']:
    person = {}
    person['name'] = item['user']['fullName'] 
    person['score_in_time'] = item['scoreInTime']
    person['score_upsolving'] = item['scoreInPractice']
    person['total'] = item['scoreSum']
    rating_57.append(person)
sorted(rating_57, key = lambda person : -person['total'])


for item in rating_89_row['rows']:
    person = {}
    person['name'] = item['user']['fullName'] 
    person['score_in_time'] = item['scoreInTime']
    person['score_upsolving'] = item['scoreInPractice']
    person['total'] = item['scoreSum']
    rating_89.append(person)
sorted(rating_89, key = lambda person : -person['total'])


# In[126]:


list_57 = merge(rating_57, kids_info)
list_89 = merge(rating_89, kids_info)


# In[ ]:





# In[127]:


splited_list_57 = split_classes(list_57, [5, 6, 7])
splited_list_89 = split_classes(list_89, [8, 9])


# In[ ]:





# In[141]:


work_book = Workbook()
sheet_5_7 = work_book.active
sheet_5_7.title = "рейтинг 5-7"
sheet_8_9 = work_book.create_sheet("рейтинг 8-9")
sheet_5 = work_book.create_sheet("рейтинг 5")
sheet_6 = work_book.create_sheet("рейтинг 6")
sheet_7 = work_book.create_sheet("рейтинг 7")
sheet_57_others = work_book.create_sheet("остальные из 5-7")

sheet_8 = work_book.create_sheet("рейтинг 8")
sheet_9 = work_book.create_sheet("рейтинг 9")
sheet_89_others = work_book.create_sheet("остальные из 8-9")


# In[142]:


head_line1 = ['имя в системе', 'всего задач', 'задачи на контестах', 'задачи в дорешку','класс',
              'почта', 'телега', 'телефон', 'город', 'школа']

important_staff1 = ['name', 'total', 'score_in_time', 'score_upsolving', 'grade', 
                  'mail', 'telegram', 'phone_number', 'city', 'school']

head_line2 = ['имя в системе', 'всего задач', 'задачи на контестах', 'задачи в дорешку', 
              'почта', 'телега', 'телефон', 'город', 'школа']
important_staff2 = ['name', 'total', 'score_in_time', 'score_upsolving', 
                  'mail', 'telegram', 'phone_number', 'city', 'school']


# In[143]:


write_to_sheet(sheet_5_7, list_57, important_staff1, head_line1)
write_to_sheet(sheet_8_9, list_89, important_staff1, head_line1)


# In[ ]:





# In[144]:


write_to_sheet(sheet_5, splited_list_57[5], important_staff2, head_line2)
write_to_sheet(sheet_6, splited_list_57[6], important_staff2, head_line2)
write_to_sheet(sheet_7, splited_list_57[7], important_staff2, head_line2)

write_to_sheet(sheet_8, splited_list_89[8], important_staff2, head_line2)
write_to_sheet(sheet_9, splited_list_89[9], important_staff2, head_line2)

write_to_sheet(sheet_57_others, splited_list_57['rest'], important_staff1, head_line1)
write_to_sheet(sheet_89_others, splited_list_89['rest'], important_staff1, head_line1)


# In[145]:


work_book.save("test.xlsx")


# In[ ]:




